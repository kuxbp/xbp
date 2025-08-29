#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from html import escape
from datetime import datetime, date, time
from openpyxl import load_workbook

# 空白判定（None か 空文字 か 空白だけ）
def _is_empty(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip() == ""
    return False

def _fmt(v) -> str:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d") if (v.hour, v.minute, v.second) == (0, 0, 0) else v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if v is None:
        return ""
    return str(v)

def excel_firstsheet_to_html(xlsx_path: Path, out_path: Path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]  # 一番左のシート

    # 値だけを2次元配列として取得
    rows = [[_fmt(v) for v in row] for row in ws.iter_rows(values_only=True)]

    if not rows:
        out_path.write_text('<table border="1" cellspacing="0" cellpadding="10">\n<tbody>\n</tbody>\n</table>\n', encoding="utf-8")
        print(f"✅ 出力: {out_path}")
        return

    # ===== 周囲の“全空”行・列をトリミング =====
    # 下端（最後の内容あり行）を探す
    last_row = -1
    for i, r in enumerate(rows):
        if any(not _is_empty(v) for v in r):
            last_row = i
    if last_row == -1:
        # 全部空
        out_path.write_text('<table border="1" cellspacing="0" cellpadding="10">\n<tbody>\n</tbody>\n</table>\n', encoding="utf-8")
        print(f"✅ 出力: {out_path}")
        return
    rows = rows[:last_row + 1]

    # 右端（最後の内容あり列）を探す
    last_col = -1
    for r in rows:
        for j, v in enumerate(r):
            if not _is_empty(v) and j > last_col:
                last_col = j
    # 各行を同じ列数に揃えつつ、右端の空列を落とす
    trimmed = []
    for r in rows:
        r = r[:last_col + 1]
        trimmed.append(r)
    rows = trimmed

    # 上端（最初の内容あり行）を探す
    first_row = 0
    while first_row < len(rows) and all(_is_empty(v) for v in rows[first_row]):
        first_row += 1
    rows = rows[first_row:]

    # 左端（最初の内容あり列）を探す
    first_col = None
    for r in rows:
        for j, v in enumerate(r):
            if not _is_empty(v):
                first_col = j if first_col is None else min(first_col, j)
    if first_col is None:
        # 内容なし
        out_path.write_text('<table border="1" cellspacing="0" cellpadding="10">\n<tbody>\n</tbody>\n</table>\n', encoding="utf-8")
        print(f"✅ 出力: {out_path}")
        return
    rows = [r[first_col:] for r in rows]

    # ===== HTML生成 =====
    # 先頭行をヘッダーとして扱う
    headers = [escape(c) for c in (rows[0] if rows else [])]
    body_rows = rows[1:] if len(rows) > 1 else []

    parts = []
    parts.append('<table border="1" cellspacing="0" cellpadding="10">')
    parts.append("<tbody>")
    if headers:
        parts.append("<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>")
    for r in body_rows:
        cells = [escape(c) for c in r]
        parts.append("<tr>" + "".join(f"<td>{c}</td>" for c in cells) + "</tr>")
    parts.append("</tbody>")
    parts.append("</table>")

    out_path.write_text("\n".join(parts) + "\n", encoding="utf-8")
    print(f"✅ 出力: {out_path}")


if __name__ == "__main__":
    input_file = Path("/Users/daisukedoyo/mygit/xbp/rekishi_dg/schedule.xlsx")
    output_file =Path("/Users/daisukedoyo/mygit/xbp/rekishi_dg/schedule.html")
    excel_firstsheet_to_html(input_file, output_file)